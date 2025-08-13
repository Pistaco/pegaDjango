import csv
import io
import os
import uuid

import PIL
import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont
import camelot
import pandas as pd
import qrcode
from barcode.writer import ImageWriter
from django.contrib.auth.models import User
from django.db.models import Exists, OuterRef, Sum, F
from django.db.models.expressions import Subquery, Value
from django.db.models.fields import IntegerField, CharField
from django.db.models.functions.comparison import Coalesce
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import render
from django.utils.timezone import now
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.reader.excel import load_workbook
from django.db import transaction
# Create your views here.
from rest_framework import viewsets, permissions, status, filters, mixins
from rest_framework.authentication import TokenAuthentication
from rest_framework.decorators import permission_classes, api_view, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError, PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView

from DjangoProject import settings
from .filters import StockFilter, FamiliaFilterSet, BodegaFilter
from .models import Cargo, Producto, Ingreso, Retiro, Usuario, StockActual, PDFUpload, ExcelUpload, Envio, EnvioDetalle, \
    Bodega, Familia, Notificacion, Pendiente, ImportRow, ImportJob
from .serializer import CargoSerializer, ProductoSerializer, IngresoSerializer, RetiroSerializer, UsuarioSerializer, \
    StockActualSerializer, ProductoStockSerializer, PDFUploadSerializer, ExcelUploadSerializer, UserSerializer, \
    EnvioSerializer, EnvioDetalleSerializer, BodegaSerializer, EnvioSerializerAnidado, FamiliaSerializer, \
    NotificacionSerializer, PendienteSerializer, ProductoInventarioSerializer, ImportRowSerializer, ImportJobSerializer, \
    ImportUploadSerializer


# permissions.py

class SoloBodeguerosVenEnviosALaBodega(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated

    def has_object_permission(self, request, view, obj):
        user = request.user
        if user.groups.filter(name="Bodeguero").exists():
            return obj.bodega_destino in user.bodegas.all()
        return True

class SoloBodeguerosVenStockDeSuBodega(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.groups.filter(name="Bodeguero").exists()


class BodegueroNOT(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.user.is_authenticated and request.user.is_active and not request.user.groups.filter(name='Bodeguero').exists():
            return True
        return False

class BodegueroYES(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.user.is_authenticated and  request.user.is_active:
            return True


def user_is_bodeguero(user):
    return user.is_authenticated and user.groups.filter(name="Bodeguero").exists()


def user_bodegas_qs(user):
    # Bodegas permitidas para el usuario
    return Bodega.objects.filter(usuarios=user)

class CurrentUserGroupView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        groups = list(request.user.groups.values_list('name', flat=True))
        print(groups)
        return Response({"groups": groups})

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name="Bodeguero").exists():
            return User.objects.filter(username=user.username).all()
        return User.objects.all()
class CargoViewSet(viewsets.ModelViewSet):
    queryset = Cargo.objects.all()
    serializer_class = CargoSerializer


class FamiliaViewSet(viewsets.ModelViewSet):
    queryset = Familia.objects.all()
    serializer_class = FamiliaSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['padre']
    filterset_class = FamiliaFilterSet
    search_fields = ['nombre']

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    ordering = ['id']
    ordering_fields = ['nombre', 'codigo_barras']
    filterset_fields = ['codigo_barras', 'nombre', 'id', 'familia']
    search_fields = ['nombre', 'codigo_barras', 'descripcion']

class NotificacionViewSet(viewsets.ModelViewSet):
    serializer_class = NotificacionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_fields = ['leido']

    def update(self, request, *args, **kwargs):
        kwargs['partial'] = True  # Permite actualizar solo algunos campos
        return super().update(request, *args, **kwargs)

    def get_queryset(self):
        if (user_is_bodeguero(self.request.user)):
            return Notificacion.objects.filter(usuario=self.request.user)
        return Notificacion.objects.all()

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)


    # views.py

def user_is_bodeguero(user):
    return user.is_authenticated and user.groups.filter(name="Bodeguero").exists()

def user_bodegas_qs(user):
    return getattr(user, 'bodegas', Bodega.objects.none()).all()

class ProductosEnMiBodegaViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated ]

    def get_queryset(self):
        user = self.request.user
        bodegas = getattr(user, "bodegas", None)
        if (not user.groups.filter(name="Bodeguero").exists()):
            return Producto.objects.all()
        if bodegas is not None:
            # Filtra productos que tienen al menos un stock en bodegas del usuario
            return Producto.objects.filter(
                Exists(
                    StockActual.objects.filter(
                        producto=OuterRef('pk'),
                        bodega__in=bodegas.all()
                    )
                )
            ).distinct()
        return Producto.objects.none()


class ProductoStockViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.select_related('stock').all()
    serializer_class = ProductoStockSerializer
    # Backends: filtrado + ordenamiento
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = StockFilter

    # Permite ordenar desde React Admin (ej: ?ordering=-cantidad)
    ordering_fields = ["id", "cantidad", "producto__nombre", "bodega__nombre"]
    ordering = ["-id"]



class ProductoEnvioViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ('id', 'nombre', 'codigo_barras')

    def get_queryset(self):
        qs = super().get_queryset()
        p = self.request.query_params
        bodega_id = p.get('bodega_id')
        con_stock = p.get('con_stock')  # "1"/"true" habilita filtro
        cantidad_min = p.get('cantidad_min')  # nuevo: umbral dinámico

        if bodega_id and str(con_stock) in ('1', 'true', 'True'):
            # Si no viene cantidad_min, asumimos 1
            try:
                cant = int(cantidad_min) if cantidad_min is not None else 1
            except ValueError:
                cant = 1

            stock_subq = StockActual.objects.filter(
                producto_id=OuterRef('pk'),
                bodega_id=bodega_id,
                cantidad__gte=cant
            )
            qs = qs.annotate(_hay_stock=Exists(stock_subq)).filter(_hay_stock=True)

        return qs

class ProductoViewSetReferenceInput(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [permissions.IsAuthenticated ]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre', 'codigo_barras']

    def get_queryset(self):
        qs = super().get_queryset()

        # a) Filtro por bodegas permitidas cuando es Bodeguero:
        if user_is_bodeguero(self.request.user):
            bodegas = user_bodegas_qs(self.request.user)
            qs = qs.filter(stock__bodega__in=bodegas).distinct()

        # b) Soporte para ReferenceInput dependiente de bodega/stock
        bodega_id = self.request.query_params.get('bodega_id')
        con_stock = self.request.query_params.get('con_stock')
        cantidad_min = self.request.query_params.get('cantidad_min')

        if bodega_id:
            qs = qs.filter(stock__bodega_id=bodega_id).distinct()
        if con_stock:
            if cantidad_min:
                qs = qs.filter(stock__cantidad__gte=cantidad_min).distinct()
            else:
                qs = qs.filter(stock__cantidad__gt=0).distinct()

        return qs


class ProductoInventarioViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Lista productos con su cantidad (stock) según bodega.
    - ?bodega_id=<id> filtra la cantidad a esa bodega.
    - Si el usuario es 'Bodeguero', solo puede consultar sus bodegas asignadas.
    """
    serializer_class = ProductoInventarioSerializer
    permission_classes = [permissions.IsAuthenticated ]



    def get_queryset(self):
        params = self.request.query_params
        bodega_id = params.get('bodega') or params.get('bodega_id')
        con_stock = params.get('con_stock') in ('1', 'true', 'True')

        # Reglas para Bodeguero
        if user_is_bodeguero(self.request.user):
            mis_bodegas = User.objects.get(username=self.request.user.username).bodegas.values_list("id")
            if not mis_bodegas:
                return Producto.objects.none()
            if bodega_id is None:
                bodega_id = next(iter(mis_bodegas))

        # Sin bodega no tiene sentido listar (evita mostrar 'cantidad' vacía)
        if bodega_id is None:
            qs = (
                Producto.objects
                .order_by('nombre')
                .distinct("nombre")
                .annotate(cantidad=F('stock__cantidad'))            # trae cantidad
            )
            return qs

        # Solo productos que TIENEN stock en esa bodega (join directo)
        qs = (
            Producto.objects
            .filter(stock__bodega_id=bodega_id)                 # exige existencia de fila stock
            .annotate(cantidad=F('stock__cantidad'))            # trae cantidad
            .order_by('nombre')
        )

        if con_stock:
            qs = qs.filter(cantidad__gt=0)

        return qs




class IngresoViewSet(viewsets.ModelViewSet):
    queryset = Ingreso.objects.all()
    serializer_class = IngresoSerializer
    permission_classes = [permissions.IsAuthenticated, BodegueroYES]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['bodega']
    search_fields = ['producto__nombre']

    def get_queryset(self):
        qs = super().get_queryset()
        if user_is_bodeguero(self.request.user):
            bodegas = user_bodegas_qs(self.request.user)
            qs = qs.filter(bodega__in=bodegas)
        return qs

class RetiroViewSet(viewsets.ModelViewSet):
    queryset = Retiro.objects.all()
    serializer_class = RetiroSerializer
    permission_classes = [permissions.IsAuthenticated, BodegueroYES ]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['bodega']
    search_fields = ['producto__nombre']

    def get_queryset(self):
        qs = super().get_queryset()
        if user_is_bodeguero(self.request.user):
            bodegas = user_bodegas_qs(self.request.user)
            qs = qs.filter(bodega__in=bodegas)
        return qs

class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

class StockViewSet(viewsets.ModelViewSet):
    queryset = StockActual.objects.all()
    serializer_class = StockActualSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]

    filterset_class = StockFilter

    # Permite ordenar desde React Admin (ej: ?ordering=-cantidad)
    ordering_fields = ["id", "cantidad", "producto__nombre", "bodega__nombre"]
    filterset_fields = ['bodega', 'producto', 'cantidad']
    search_fields = ['nombre', 'codigo_barras', 'descripcion']

class StockBajoStockViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StockActualSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = StockFilter
    ordering_fields = ["id", "cantidad", "producto__nombre", "bodega__nombre"]
    filterset_fields = ['bodega', 'producto', 'cantidad']
    search_fields = ['nombre', 'codigo_barras', 'descripcion']

    def get_queryset(self):
        return StockActual.objects.filter(cantidad__lt=5)

class StockDeMiBodegaViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StockActualSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = StockFilter
    ordering_fields = ["id", "cantidad", "producto__nombre", "bodega__nombre"]
    filterset_fields = ['bodega', 'producto', 'cantidad']
    search_fields = ['nombre', 'codigo_barras', 'descripcion']

    def get_queryset(self):
        user = self.request.user
        bodegas = getattr(user, "bodegas", None)
        if bodegas is not None:
            return StockActual.objects.filter(bodega__in=bodegas.all())
        return StockActual.objects.none()

class StockDeMiBodegaBajoStockViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StockActualSerializer
    permission_classes = [IsAuthenticated ]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = StockFilter
    ordering_fields = ["id", "cantidad", "producto__nombre", "bodega__nombre"]
    filterset_fields = ['bodega', 'producto', 'cantidad']
    search_fields = ['nombre', 'codigo_barras', 'descripcion']

    def get_queryset(self):
        user = self.request.user
        bodegas = getattr(user, "bodegas", None)
        if bodegas is not None:
            return StockActual.objects.filter(bodega__in=bodegas.all()).filter(cantidad__lt=5)
        return StockActual.objects.none()
class EnvioViewSet(viewsets.ModelViewSet):
    queryset = Envio.objects.all()
    serializer_class = EnvioSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['confirmado', 'bodega_origen', 'bodega_destino']

class EnvioAnidadoViewSet(viewsets.ModelViewSet):
    queryset = Envio.objects.all().prefetch_related('detalles')
    serializer_class = EnvioSerializerAnidado
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['confirmado', 'bodega_origen', 'bodega_destino', 'usuario']
    name = 'EnvioAnidado'

    def get_queryset(self):
        user = self.request.user
        if user.groups.filter(name="Bodeguero").exists():
            return Envio.objects.filter(bodega_destino__in=user.bodegas.all())
        return Envio.objects.all()


    def perform_create(self, serializer):
        # Asigna automáticamente el usuario autenticado
        serializer.save(usuario=self.request.user)

    def perform_update(self, serializer):
        # Evita que alguien cambie el "usuario" del envío por PATCH/PUT
        serializer.save(usuario=self.get_object().usuario)

class EnvioEnProgresoViewSet(viewsets.ModelViewSet):
    """
    Lista/consulta envíos en progreso (confirmado=False) del usuario autenticado.
    """
    queryset = Envio.objects.all().prefetch_related('detalles')
    serializer_class = EnvioSerializerAnidado
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['confirmado', 'bodega_origen', 'bodega_destino', 'usuario']

    def get_queryset(self):
        user = self.request.user
        return (Envio.objects
                .select_related('bodega_origen', 'bodega_destino', 'usuario')
                .filter(usuario_id=user.id, confirmado=False)
                )


class EnvioRecibidosViewSet(viewsets.ModelViewSet):
    """
    Lista/consulta envíos en progreso (confirmado=False) del usuario autenticado.
    """
    queryset = Envio.objects.all().prefetch_related('detalles')
    serializer_class = EnvioSerializerAnidado
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['confirmado', 'bodega_origen', 'bodega_destino', 'usuario']

    def get_queryset(self):
        user = self.request.user
        return (Envio.objects
                .select_related('bodega_origen', 'bodega_destino', 'usuario')
                .filter(bodega_destino__in=user.bodegas.all())
                )


class EnvioDetalleViewSet(viewsets.ModelViewSet):
    queryset = EnvioDetalle.objects.all()
    serializer_class = EnvioDetalleSerializer

class BodegaViewSet(viewsets.ModelViewSet):
    queryset = Bodega.objects.all()
    serializer_class = BodegaSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre']
    filterset_class = BodegaFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        if user_is_bodeguero(self.request.user):
            qs = user_bodegas_qs(self.request.user)
        return qs


class BodegaTodasViewSet(viewsets.ModelViewSet):
    queryset = Bodega.objects.all()
    serializer_class = BodegaSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['nombre']



class PendienteViewSet(viewsets.ModelViewSet):
    queryset = Pendiente.objects.all().select_related('producto')
    serializer_class = PendienteSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['producto', 'completado']
    search_fields = ['descripcion']

class PDFUploadView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, BodegueroNOT]

    def post(self, request):
        serializer = PDFUploadSerializer(data=request.data)
        if serializer.is_valid():
            pdf_obj = serializer.save()
            pdf_path = pdf_obj.archivo.path

            try:
                tables = camelot.read_pdf(pdf_path, pages='all')

                for table in tables:
                    df = table.df  # pandas DataFrame

                    # Ejemplo: asumir primera fila es encabezado
                    headers = df.iloc[0]
                    data_rows = df.iloc[1:]

                    for row in data_rows.itertuples(index=False):
                        row_dict = dict(zip(headers, row))

                        if row_dict.get("DETALLE") == "":
                            continue  # Salta esta fila
                        producto = Producto.objects.create(
                            nombre=row_dict.get("DETALLE"),
                            descripcion=row_dict.get("Descripción"),
                            codigo_barras=row_dict.get("Código de barras") or str(uuid.uuid4())[:20]
                        )
                        stock = StockActual.objects.get(producto_id=producto)
                        stock.cantidad = int(row_dict.get("CANTIDAD", 0))
                        stock.save()


                response = PDFUploadSerializer(pdf_obj)
                pdf_obj.delete()
                return Response(response.data, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ExcelUploadView(APIView):

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, BodegueroNOT]

    def post(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_obj = serializer.save()
            archivo = serializer.validated_data['archivo']
            wb = load_workbook(filename=archivo, data_only=True)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                if not row_data.get("Nombre"):
                    continue

                producto = Producto.objects.create(
                    nombre=row_data.get("Nombre"),
                    descripcion=row_data.get("Descripción", ""),
                    codigo_barras=row_data.get("Código de barras") or str(uuid.uuid4())[:20],
                    centro_costo=row_data.get("Centro Costo")
                )



            response = ExcelUploadSerializer(excel_obj)
            excel_obj.delete()
            return Response(response.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Función auxiliar para agregar texto arriba de una imagen
def add_product_name_to_image(img: Image.Image, text: str) -> Image.Image:
    if img.mode != "RGB":
        img = img.convert("RGB")

    font_path = os.path.join(os.path.dirname(__file__), "fonts", "FreeMono", "FreeMonospaced.ttf")
    font = ImageFont.truetype(font_path, size=40)
    # Medidas
    bbox = font.getbbox(text)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]

    img_width, img_height = img.size

    # Nueva imagen con espacio extra arriba
    new_height = img_height + text_height + 10
    new_img = Image.new("RGB", (img_width, new_height), "white")

    # Dibujar texto
    draw = ImageDraw.Draw(new_img)
    text_x = (img_width - text_width) // 2
    draw.text((text_x, 5), text, font=font, fill="black")

    # Pegar la imagen original abajo
    new_img.paste(img, (0, text_height + 10))

    return new_img


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def barcode_image_on_demand(request, pk):
    try:
        producto = Producto.objects.get(pk=pk)
    except Producto.DoesNotExist:
        return HttpResponse("Producto no encontrado", status=404)

    # Generar código de barras
    barcode_class = barcode.get_barcode_class('code128')
    buffer = io.BytesIO()
    barcode_class(producto.codigo_barras, writer=ImageWriter()).write(buffer)

    # Abrir imagen y agregar nombre
    buffer.seek(0)
    img = Image.open(buffer)
    img_with_name = add_product_name_to_image(img, producto.nombre)

    # Guardar a buffer y enviar
    output = io.BytesIO()
    img_with_name.save(output, format="PNG")
    output.seek(0)

    return HttpResponse(
        output.getvalue(),
        content_type="image/png",
        headers={
            'Content-Disposition': f'attachment; filename="{producto.codigo_barras}.png"'
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def sqr_image_on_demand(request, pk):
    try:
        producto = Producto.objects.get(pk=pk)
    except Producto.DoesNotExist:
        return HttpResponse("Producto no encontrado", status=404)

    # Generar código QR
    qr = qrcode.make(producto.codigo_barras)
    img_with_name = add_product_name_to_image(qr, producto.nombre)

    # Guardar a buffer y enviar
    output = io.BytesIO()
    img_with_name.save(output, format="PNG")
    output.seek(0)

    return HttpResponse(
        output.getvalue(),
        content_type="image/png",
        headers={
            'Content-Disposition': f'attachment; filename="{producto.codigo_barras}.png"'
        }
    )


class ImportJobViewSet(viewsets.ModelViewSet):
    """
    /api/importaciones/
    - POST /upload/ (multipart): bodega, file
    - GET listar trabajos (limitado por rol)
    - GET /{id}/rows/ para ver detalle
    """
    queryset = ImportJob.objects.all().order_by('-created_at')
    serializer_class = ImportJobSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user_is_bodeguero(user):
            # Solo trabajos en bodegas a las que tiene acceso:
            # Si tienes una tabla bodega_usuarios, filtra por ella. Aquí asumimos relación directa.
            return qs.filter(usuario=user)
        return qs

    def perform_create(self, serializer):
        # No se usa (creamos via /upload/)
        serializer.save(usuario=self.request.user)

    @action(detail=False, methods=['post'], url_path='upload')
    def upload(self, request):
        """
        Carga un archivo Excel/CSV con columnas: nombre, cantidad, precio
        Requiere bodega (id).
        """
        ser = ImportUploadSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        bodega = ser.validated_data['bodega']
        familia = ser.validated_data['familia']
        f = ser.validated_data['file']

        # Si es Bodeguero, restringir bodega
        if user_is_bodeguero(request.user):
            # Aquí implementar tu lógica real (p.ej., validar que pertenece a esa bodega)
            # raise PermissionDenied si no corresponde
            pass

        job = ImportJob.objects.create(
            usuario=request.user,
            bodega=bodega,
            filename=f.name,
            status='processing'
        )

        try:
            # Parse archivo
            rows = []
            if f.name.lower().endswith('.csv'):
                text = f.read().decode('utf-8', errors='ignore').splitlines()
                reader = csv.DictReader(text)
                for i, r in enumerate(reader, start=2):  # fila 2 por encabezado
                    rows.append({
                        'row_number': i,
                        'nombre': (r.get('nombre') or '').strip(),
                        'cantidad': (r.get('cantidad') or '').strip(),
                        'precio': (r.get('precio') or '').strip(),
                    })
            else:
                wb = load_workbook(filename=f, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.rows)]
                # esperamos columnas: nombre, cantidad, precio
                colmap = {h: idx for idx, h in enumerate(headers)}
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    def get(h):
                        idx = colmap.get(h, None)
                        return (row[idx].value if idx is not None else None)
                    rows.append({
                        'row_number': i,
                        'nombre': (get('nombre') or '').strip() if get('nombre') else '',
                        'cantidad': get('cantidad'),
                        'precio': get('precio'),
                    })

            to_create_rows = []
            ok_count = 0

            with transaction.atomic():
                for r in rows:
                    nombre = r['nombre']
                    cantidad = r['cantidad']
                    precio = r['precio']
                    if nombre == '' or cantidad is None or precio is None:
                        to_create_rows.append(ImportRow(
                            import_job=job,
                            row_number=r['row_number'],
                            nombre=nombre or '',
                            cantidad=cantidad or 0,
                            precio=precio or 0,
                            status='error',
                            message='Fila incompleta: nombre/cantidad/precio'
                        ))
                        continue

                    try:
                        cantidad_val = float(cantidad)
                        precio_val = float(precio)
                        if cantidad_val < 0 or precio_val < 0:
                            raise ValueError('Valores negativos no permitidos')

                        # Producto por nombre (puedes cambiar a case-insensitive unique si quieres)
                        codigo = uuid.uuid4()
                        codigo = str(codigo)[:20]
                        producto, created = Producto.objects.get_or_create(
                            nombre=nombre,
                            defaults={'precio': precio_val, 'familia': familia, 'codigo_barras': codigo},

                        )
                        if not created:
                            # actualiza precio si distinto
                            if producto.precio != precio_val:
                                producto.precio = precio_val
                                producto.save(update_fields=['precio'])

                        # Upsert de stock: REEMPLAZA cantidad por la del Excel
                        stock, s_created = StockActual.objects.get_or_create(
                            producto=producto, bodega=bodega,
                            defaults={'cantidad': cantidad_val}
                        )
                        if not s_created:
                            # Para “sumar”, usa: stock.cantidad = stock.cantidad + Decimal(cantidad_val)
                            stock.cantidad = cantidad_val
                            stock.save(update_fields=['cantidad'])

                        to_create_rows.append(ImportRow(
                            import_job=job,
                            row_number=r['row_number'],
                            nombre=nombre,
                            cantidad=cantidad_val,
                            precio=precio_val,
                            producto=producto,
                            status='ok',
                            message='OK' if created or s_created else 'Actualizado'
                        ))
                        ok_count += 1

                    except Exception as e:
                        to_create_rows.append(ImportRow(
                            import_job=job,
                            row_number=r['row_number'],
                            nombre=nombre or '',
                            cantidad=float(r['cantidad']) if r['cantidad'] not in (None, '') else 0,
                            precio=float(r['precio']) if r['precio'] not in (None, '') else 0,
                            status='error',
                            message=str(e)[:500]
                        ))

                ImportRow.objects.bulk_create(to_create_rows, batch_size=500)
                job.total_rows = len(rows)
                job.status = 'done'
                job.finished_at = now()
                job.save(update_fields=['total_rows', 'status', 'finished_at'])

            return Response(ImportJobSerializer(job).data, status=status.HTTP_201_CREATED)

        except Exception as e:
            job.status = 'failed'
            job.error_message = str(e)[:2000]
            job.finished_at = now()
            job.save(update_fields=['status', 'error_message', 'finished_at'])
            return Response({'detail': 'Fallo la importación', 'error': str(e)}, status=400)

    @action(detail=True, methods=['get'], url_path='rows')
    def rows(self, request, pk=None):
        job = self.get_object()
        rows = job.rows.all().order_by('row_number')
        return Response(ImportRowSerializer(rows, many=True).data)

class ImportRowViewSet(viewsets.ModelViewSet):
    """
    /api/import_rows/?import_job=<ID>&status=<ok|error|skipped>
    """
    queryset = ImportRow.objects.select_related('import_job', 'producto').order_by('row_number')
    serializer_class = ImportRowSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['import_job', 'status']

    def get_queryset(self):
        qs = super().get_queryset()
        # Bodeguero solo ve sus filas (vía el job)
        if user_is_bodeguero(self.request.user):
            qs = qs.filter(import_job__usuario=self.request.user)
        return qs